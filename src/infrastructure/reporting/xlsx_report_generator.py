"""Informe de sesión en Excel (.xlsx) vía openpyxl."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from ...domain.models import UsageRecord, human_bytes
from ...domain.ports import ReportGenerator


class XlsxReportGenerator(ReportGenerator):
    def __init__(self, output_dir: str | Path | None = None) -> None:
        self._dir = Path(output_dir or (Path.home() / "Desktop"))
        self._dir.mkdir(parents=True, exist_ok=True)

    def generate(
        self, records: Iterable[UsageRecord], destination_hint: str | None = None
    ) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = self._dir / (destination_hint or f"NetLeak_report_{stamp}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Consumo"
        headers = ["Aplicación", "Subida", "Bajada", "Total",
                   "Subida (B)", "Bajada (B)"]
        ws.append(headers)
        head_fill = PatternFill("solid", fgColor="3B82F6")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = head_fill

        rows = sorted(records, key=lambda x: -(x.bytes_sent + x.bytes_recv))
        tot_s = tot_r = 0
        for r in rows:
            total = r.bytes_sent + r.bytes_recv
            ws.append([r.app_name, human_bytes(r.bytes_sent), human_bytes(r.bytes_recv),
                       human_bytes(total), r.bytes_sent, r.bytes_recv])
            tot_s += r.bytes_sent
            tot_r += r.bytes_recv
        ws.append([])
        total_row = ["TOTAL", human_bytes(tot_s), human_bytes(tot_r),
                     human_bytes(tot_s + tot_r), tot_s, tot_r]
        ws.append(total_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        widths = [34, 12, 12, 12, 14, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        wb.save(path)
        return str(path)
